"""Import en masse via Excel/CSV — 100% local, sans IA.

Format Excel : un classeur avec une feuille par type d'entité, dont les
colonnes sont décrites dans TEMPLATE_SHEETS. Une feuille "_Notice" en tête
documente l'usage pour l'humain et pour Claude (quand l'utilisateur le
demande de pré-remplir le fichier depuis l'extérieur).

Format CSV : une seule entité à la fois, le type est fourni par l'appelant
(query param).

Pipeline :
  1. parse_workbook(file) → ParsedImport (rows par type avec status)
  2. l'utilisateur peut résoudre les ambiguïtés via le frontend (dropdown
     compte, etc.) ; les résolutions sont renvoyées au commit
  3. commit_parsed(parsed, resolutions, user) crée/met à jour les entités
     dans un ImportBatch unique → undo possible
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import date as DateType, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import (
    Account, AccountType, AutoSaving, Charge, Frequency, ImportBatch,
    ImportedEntity, Income, IncomeType, OneTimeTransfer, PaymentMethod,
    Purchase, RecurringTransfer, SplitMode, User,
)
from services.access import user_can_write_account

logger = logging.getLogger(__name__)


# ============================================================
# Schéma du template
# ============================================================

class Col:
    """Définition d'une colonne du template."""
    def __init__(self, key: str, label: str, required: bool = False,
                 example: str = "", hint: str = ""):
        self.key = key
        self.label = label
        self.required = required
        self.example = example
        self.hint = hint


TEMPLATE_SHEETS: dict[str, list[Col]] = {
    "Comptes": [
        Col("name", "Nom du compte", required=True, example="Compte courant BNP"),
        Col("bank", "Banque", required=True, example="BNP Paribas"),
        Col("type", "Type", required=True, example="Compte courant",
            hint="Compte courant, Livret A, LDDS, LEP, PEL, CEL, PEA, Assurance vie, Compte joint, Compte épargne, Compte titres, Autre"),
        Col("initial_balance", "Solde initial (€)", example="1500.00"),
        Col("space", "Espace", example="perso", hint="perso ou pro"),
        Col("notes", "Notes", example=""),
    ],
    "Revenus": [
        Col("source", "Source", required=True, example="Salaire ACME"),
        Col("amount", "Montant (€)", required=True, example="2500.00"),
        Col("day_of_month", "Jour du mois", required=True, example="2"),
        Col("type", "Type", example="Régulier", hint="Régulier, Ponctuel, Variable"),
        Col("account_name", "Compte cible", required=True, example="Compte courant BNP"),
        Col("valid_from", "Date début (YYYY-MM-DD)", example=""),
        Col("valid_to", "Date fin (YYYY-MM-DD)", example=""),
    ],
    "Charges": [
        Col("label", "Libellé", required=True, example="Loyer"),
        Col("total_amount", "Montant total (€)", required=True, example="850.00"),
        Col("day_of_month", "Jour du mois", required=True, example="5"),
        Col("frequency", "Fréquence", example="Mensuelle",
            hint="Mensuelle, Bimensuelle, Trimestrielle, Semestrielle, Annuelle"),
        Col("split_mode", "Mode partage", example="Perso",
            hint="Perso, Égal, Pourcentage, Montant fixe, Par utilisateur"),
        Col("num_colocs", "Nb colocs (si Égal)", example="1"),
        Col("account_name", "Compte cible", required=True, example="Compte courant BNP"),
        Col("valid_from", "Date début", example=""),
        Col("valid_to", "Date fin", example=""),
        Col("notes", "Notes", example=""),
    ],
    "Virements récurrents": [
        Col("label", "Libellé", required=True, example="Vir. épargne"),
        Col("amount", "Montant (€)", required=True, example="200.00"),
        Col("source_account_name", "Compte source", required=True, example="Compte courant BNP"),
        Col("dest_account_name", "Compte destination", required=True, example="Livret A"),
        Col("day_of_month", "Jour du mois", required=True, example="10"),
        Col("frequency", "Fréquence", example="Mensuelle"),
        Col("valid_from", "Date début", example=""),
        Col("valid_to", "Date fin", example=""),
    ],
    "Virements ponctuels": [
        Col("date", "Date (YYYY-MM-DD)", required=True, example="2026-05-12"),
        Col("label", "Libellé", required=True, example="Remboursement Naïm"),
        Col("amount", "Montant (€)", required=True, example="42.50"),
        Col("source_account_name", "Compte source", required=True, example="Compte courant BNP"),
        Col("dest_account_name", "Compte destination", required=True, example="Livret A"),
    ],
    "Épargne": [
        Col("label", "Libellé", required=True, example="Mise de côté"),
        Col("amount", "Montant (€)", required=True, example="100.00"),
        Col("source_account_name", "Compte source", required=True, example="Compte courant BNP"),
        Col("dest_account_name", "Compte destination", required=True, example="Livret A"),
        Col("day_of_month", "Jour du mois", required=True, example="1"),
        Col("valid_from", "Date début", example=""),
        Col("valid_to", "Date fin", example=""),
    ],
    "Achats": [
        Col("date", "Date (YYYY-MM-DD)", required=True, example="2026-05-12"),
        Col("description", "Description", required=True, example="Carrefour"),
        Col("total_amount", "Montant (€)", required=True, example="47.32"),
        Col("nb_installments", "Étalement", example="1", hint="1 = comptant, 3 = 3× sans frais, etc."),
        Col("category", "Catégorie", example="Alimentation"),
        Col("payment_method", "Moyen de paiement", example="CB",
            hint="CB, Virement, Espèces, Chèque, Prélèvement, Autre"),
        Col("account_name", "Compte cible", required=True, example="Compte courant BNP"),
    ],
}


# ============================================================
# Génération du template Excel
# ============================================================

def generate_template() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet "_Notice" en tête (visible en premier)
    notice = wb.create_sheet("_Notice", 0)
    _write_notice(notice)

    for sheet_name, cols in TEMPLATE_SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        _write_sheet_header(ws, cols)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_notice(ws) -> None:
    ws.column_dimensions["A"].width = 100
    rows = [
        ("Budget Multi-Coloc — Template d'import en masse", True),
        ("", False),
        ("Comment ça marche :", True),
        ("1. Remplis une ou plusieurs feuilles ci-après selon ce que tu veux importer.", False),
        ("2. Tu peux supprimer les lignes d'exemple si tu n'en as pas besoin.", False),
        ("3. Le 'Compte cible' / 'Compte source' DOIT correspondre au nom EXACT", False),
        ("   d'un compte existant OU être saisi dans la feuille 'Comptes' (les comptes", False),
        ("   seront créés en premier au commit).", False),
        ("4. Dates au format YYYY-MM-DD (ex: 2026-05-12).", False),
        ("5. Montants en euros, point ou virgule décimale acceptés (ex: 1234.56).", False),
        ("6. Les colonnes d'en-tête doivent rester telles quelles — ne renomme pas.", False),
        ("", False),
        ("Tu peux donner ce template à un LLM externe (ChatGPT, Claude.ai, Gemini)", False),
        ("avec ton extrait bancaire pour qu'il te le remplisse automatiquement.", False),
        ("Le LLM doit garder les noms de colonnes et de feuilles inchangés.", False),
    ]
    for i, (text, bold) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12)


def _write_sheet_header(ws, cols: list[Col]) -> None:
    header_fill = PatternFill("solid", fgColor="E8E1D3")
    header_font = Font(bold=True)
    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=col.label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(i)].width = max(16, len(col.label) + 2)

    # Ligne d'exemple (en italique, gris)
    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=i, value=col.example)
        cell.font = Font(italic=True, color="888888")

    # Hint sur la ligne 3 quand renseigné
    has_hint = any(c.hint for c in cols)
    if has_hint:
        for i, col in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=i, value=col.hint)
            cell.font = Font(italic=True, color="AAAAAA", size=10)


# ============================================================
# Parsing
# ============================================================

class ParsedRow:
    """Une ligne parsée du fichier d'import."""
    def __init__(self, idx: int, data: dict[str, Any]):
        self.idx = idx
        self.data = data
        self.status: str = "ready"  # 'ready' | 'ambiguous' | 'error' | 'skipped'
        self.issues: list[str] = []
        self.suggestions: dict[str, list[str]] = {}

    def to_dict(self) -> dict:
        return {
            "idx": self.idx,
            "data": self.data,
            "status": self.status,
            "issues": self.issues,
            "suggestions": self.suggestions,
        }


class ParsedImport:
    def __init__(self):
        self.sheets: dict[str, list[ParsedRow]] = {}

    def to_dict(self) -> dict:
        return {
            name: [r.to_dict() for r in rows]
            for name, rows in self.sheets.items()
        }

    @classmethod
    def from_dict(cls, data: dict) -> "ParsedImport":
        p = cls()
        for name, rows in data.items():
            parsed_rows = []
            for r in rows:
                pr = ParsedRow(r["idx"], r["data"])
                pr.status = r.get("status", "ready")
                pr.issues = r.get("issues", [])
                pr.suggestions = r.get("suggestions", {})
                parsed_rows.append(pr)
            p.sheets[name] = parsed_rows
        return p


def parse_workbook(file_bytes: bytes, filename: str) -> ParsedImport:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        # CSV impose un seul type → on lit la première colonne pour deviner
        return _parse_csv(file_bytes)
    return _parse_xlsx(file_bytes)


def _parse_xlsx(file_bytes: bytes) -> ParsedImport:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = ParsedImport()
    for sheet_name, cols in TEMPLATE_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        result.sheets[sheet_name] = _parse_sheet(ws, cols)
    return result


def _parse_sheet(ws, cols: list[Col]) -> list[ParsedRow]:
    headers_row = [
        (ws.cell(row=1, column=i + 1).value or "").strip()
        for i in range(len(cols))
    ]
    # Mapping label → col.key
    label_to_key = {c.label: c.key for c in cols}
    key_by_idx = [label_to_key.get(h, None) for h in headers_row]

    rows: list[ParsedRow] = []
    for r in range(2, ws.max_row + 1):
        # Ignorer la ligne d'exemple (italique) si toutes les valeurs matchent
        # les exemples — heuristique simple : on tente de parser, si tout
        # match l'exemple textuellement on skip
        raw = {}
        any_value = False
        for ci, key in enumerate(key_by_idx):
            if not key:
                continue
            v = ws.cell(row=r, column=ci + 1).value
            if v is not None and str(v).strip() != "":
                any_value = True
            raw[key] = v
        if not any_value:
            continue
        # Skip les lignes d'exemple : compare au texte d'exemple
        if r == 2:
            example_match = all(
                str(raw.get(c.key, "") or "").strip() == c.example
                for c in cols if c.example
            )
            if example_match:
                continue
        # Ligne 3 = hints (italique) si présents, skip aussi
        if r == 3 and any(c.hint for c in cols):
            hint_match = any(
                str(raw.get(c.key, "") or "").strip() == c.hint
                for c in cols if c.hint
            )
            if hint_match:
                continue
        row = ParsedRow(r, _normalize_row(raw, cols))
        rows.append(row)
    return rows


def _parse_csv(file_bytes: bytes) -> ParsedImport:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    # Détection du type de CSV : on cherche le sheet dont les labels matchent le plus
    best_sheet = None
    best_score = 0
    for sheet_name, cols in TEMPLATE_SHEETS.items():
        labels = {c.label for c in cols}
        score = sum(1 for h in headers if h in labels)
        if score > best_score:
            best_score = score
            best_sheet = sheet_name
    result = ParsedImport()
    if not best_sheet or best_score < 2:
        return result  # CSV non reconnu
    cols = TEMPLATE_SHEETS[best_sheet]
    label_to_key = {c.label: c.key for c in cols}
    rows = []
    for i, raw in enumerate(reader, start=2):
        mapped = {label_to_key[k]: v for k, v in raw.items() if k in label_to_key}
        if not any(str(v or "").strip() for v in mapped.values()):
            continue
        rows.append(ParsedRow(i, _normalize_row(mapped, cols)))
    result.sheets[best_sheet] = rows
    return result


def _normalize_row(raw: dict, cols: list[Col]) -> dict:
    """Normalise les types : strip strings, convertit dates et nombres."""
    out: dict[str, Any] = {}
    for c in cols:
        v = raw.get(c.key)
        if v is None:
            out[c.key] = None
            continue
        if isinstance(v, (datetime, DateType)):
            out[c.key] = v.isoformat()[:10] if "date" in c.key or c.key.startswith("valid_") else str(v)
            continue
        s = str(v).strip()
        if s == "":
            out[c.key] = None
            continue
        out[c.key] = s
    return out


# ============================================================
# Validation
# ============================================================

def validate(parsed: ParsedImport, db: Session, user: User) -> ParsedImport:
    """Marque chaque ligne ready/ambiguous/error selon les données existantes.

    Important : les comptes définis dans la feuille 'Comptes' du fichier sont
    considérés comme "à créer" et donc résolvables pour les autres feuilles.
    """
    # Index des comptes existants accessibles
    existing_accounts = {
        a.name.strip().lower(): a
        for a in db.query(Account).all()
    }
    # Comptes qui seront créés par cet import
    pending_account_names = set()
    if "Comptes" in parsed.sheets:
        for row in parsed.sheets["Comptes"]:
            _validate_account_row(row)
            if row.status != "error":
                name_l = (row.data.get("name") or "").strip().lower()
                if name_l and name_l not in existing_accounts:
                    pending_account_names.add(name_l)

    # Comptes effectivement disponibles après commit
    available_names = set(existing_accounts.keys()) | pending_account_names

    validators = {
        "Revenus": (_validate_income_row, ["account_name"]),
        "Charges": (_validate_charge_row, ["account_name"]),
        "Virements récurrents": (_validate_rec_transfer_row, ["source_account_name", "dest_account_name"]),
        "Virements ponctuels": (_validate_ot_transfer_row, ["source_account_name", "dest_account_name"]),
        "Épargne": (_validate_saving_row, ["source_account_name", "dest_account_name"]),
        "Achats": (_validate_purchase_row, ["account_name"]),
    }
    for sheet_name, (validator, account_fields) in validators.items():
        if sheet_name not in parsed.sheets:
            continue
        for row in parsed.sheets[sheet_name]:
            validator(row)
            # Résolution des comptes référencés
            for field in account_fields:
                ref = (row.data.get(field) or "").strip()
                if not ref:
                    row.issues.append(f"Compte manquant ({field})")
                    row.status = "error"
                    continue
                if ref.lower() not in available_names:
                    row.issues.append(f"Compte « {ref} » introuvable")
                    row.suggestions[field] = sorted(
                        [a.name for a in existing_accounts.values()]
                    )
                    if row.status == "ready":
                        row.status = "ambiguous"
    return parsed


def _safe_decimal(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", ".").replace(" ", ""))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _safe_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _safe_date(v: Any) -> Optional[DateType]:
    if v is None or v == "":
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _validate_account_row(row: ParsedRow) -> None:
    if not row.data.get("name"):
        row.status = "error"; row.issues.append("Nom du compte manquant")
    if not row.data.get("bank"):
        row.status = "error"; row.issues.append("Banque manquante")
    type_str = (row.data.get("type") or "").strip()
    if not type_str:
        row.status = "error"; row.issues.append("Type manquant")
    else:
        try:
            AccountType(type_str)
        except ValueError:
            row.status = "ambiguous"
            row.issues.append(f"Type inconnu : {type_str}")
            row.suggestions["type"] = [t.value for t in AccountType]
    if row.data.get("initial_balance") and _safe_decimal(row.data["initial_balance"]) is None:
        row.status = "error"; row.issues.append("Solde initial invalide")


def _validate_income_row(row: ParsedRow) -> None:
    if not row.data.get("source"):
        row.status = "error"; row.issues.append("Source manquante")
    if _safe_decimal(row.data.get("amount")) is None:
        row.status = "error"; row.issues.append("Montant invalide")
    d = _safe_int(row.data.get("day_of_month"))
    if d is None or not (1 <= d <= 31):
        row.status = "error"; row.issues.append("Jour du mois invalide (1-31)")
    if row.data.get("type"):
        try:
            IncomeType(row.data["type"])
        except ValueError:
            row.status = "ambiguous"; row.issues.append(f"Type revenu inconnu : {row.data['type']}")
            row.suggestions["type"] = [t.value for t in IncomeType]


def _validate_charge_row(row: ParsedRow) -> None:
    if not row.data.get("label"):
        row.status = "error"; row.issues.append("Libellé manquant")
    if _safe_decimal(row.data.get("total_amount")) is None:
        row.status = "error"; row.issues.append("Montant invalide")
    d = _safe_int(row.data.get("day_of_month"))
    if d is None or not (1 <= d <= 31):
        row.status = "error"; row.issues.append("Jour du mois invalide")
    if row.data.get("frequency"):
        try:
            Frequency(row.data["frequency"])
        except ValueError:
            row.status = "ambiguous"; row.issues.append(f"Fréquence inconnue : {row.data['frequency']}")
            row.suggestions["frequency"] = [f.value for f in Frequency]
    if row.data.get("split_mode"):
        try:
            SplitMode(row.data["split_mode"])
        except ValueError:
            row.status = "ambiguous"; row.issues.append(f"Mode partage inconnu : {row.data['split_mode']}")
            row.suggestions["split_mode"] = [s.value for s in SplitMode]


def _validate_rec_transfer_row(row: ParsedRow) -> None:
    if not row.data.get("label"):
        row.status = "error"; row.issues.append("Libellé manquant")
    if _safe_decimal(row.data.get("amount")) is None:
        row.status = "error"; row.issues.append("Montant invalide")
    d = _safe_int(row.data.get("day_of_month"))
    if d is None or not (1 <= d <= 31):
        row.status = "error"; row.issues.append("Jour du mois invalide")


def _validate_ot_transfer_row(row: ParsedRow) -> None:
    if not row.data.get("label"):
        row.status = "error"; row.issues.append("Libellé manquant")
    if _safe_decimal(row.data.get("amount")) is None:
        row.status = "error"; row.issues.append("Montant invalide")
    if _safe_date(row.data.get("date")) is None:
        row.status = "error"; row.issues.append("Date invalide")


def _validate_saving_row(row: ParsedRow) -> None:
    _validate_rec_transfer_row(row)


def _validate_purchase_row(row: ParsedRow) -> None:
    if not row.data.get("description"):
        row.status = "error"; row.issues.append("Description manquante")
    if _safe_decimal(row.data.get("total_amount")) is None:
        row.status = "error"; row.issues.append("Montant invalide")
    if _safe_date(row.data.get("date")) is None:
        row.status = "error"; row.issues.append("Date invalide")


# ============================================================
# Commit
# ============================================================

def commit_parsed(
    db: Session,
    user: User,
    parsed: ParsedImport,
    resolutions: dict,
) -> ImportBatch:
    """Applique toutes les lignes 'ready' (et les 'ambiguous' avec résolutions
    fournies). resolutions est un dict {sheet_name: {row_idx: {field: value}}}
    pour patcher les valeurs ambigües avant insertion.

    Crée un ImportBatch unique pour permettre l'undo global.
    """
    batch = ImportBatch(
        user_id=user.id,
        source_type="bulk_excel",
        raw_response=None,
        summary="(import en masse)",
        status="committed",
    )
    db.add(batch)
    db.flush()

    counters: dict[str, int] = {}

    def _bump(kind: str) -> None:
        counters[kind] = counters.get(kind, 0) + 1

    # Étape 1 : créer les comptes en premier (les autres lignes en dépendent)
    name_to_account: dict[str, Account] = {
        a.name.strip().lower(): a for a in db.query(Account).all()
    }

    if "Comptes" in parsed.sheets:
        for row in parsed.sheets["Comptes"]:
            data = _apply_resolutions(row, resolutions.get("Comptes"))
            if row.status == "error":
                continue
            name = (data.get("name") or "").strip()
            if not name:
                continue
            key = name.lower()
            if key in name_to_account:
                continue  # déjà existant — pas d'écrasement
            try:
                acc_type = AccountType(data.get("type") or "Compte courant")
            except ValueError:
                acc_type = AccountType.COURANT
            space = (data.get("space") or "perso").strip()
            if space not in ("perso", "pro"):
                space = "perso"
            acc = Account(
                user_id=user.id,
                name=name,
                bank=(data.get("bank") or "").strip(),
                type=acc_type,
                initial_balance=_safe_decimal(data.get("initial_balance")) or Decimal(0),
                notes=data.get("notes") or None,
                space=space,
            )
            db.add(acc)
            db.flush()
            db.add(ImportedEntity(batch_id=batch.id, entity_type="account", entity_id=acc.id))
            name_to_account[key] = acc
            _bump("comptes")

    def _resolve_account(name_raw: Optional[str]) -> Optional[Account]:
        if not name_raw:
            return None
        return name_to_account.get(name_raw.strip().lower())

    # Étape 2 : autres entités
    sheet_processors = {
        "Revenus": _process_income,
        "Charges": _process_charge,
        "Virements récurrents": _process_rec_transfer,
        "Virements ponctuels": _process_ot_transfer,
        "Épargne": _process_saving,
        "Achats": _process_purchase,
    }
    for sheet_name, processor in sheet_processors.items():
        if sheet_name not in parsed.sheets:
            continue
        for row in parsed.sheets[sheet_name]:
            data = _apply_resolutions(row, resolutions.get(sheet_name))
            if row.status == "error":
                continue
            try:
                entity = processor(db, user, data, _resolve_account)
            except Exception as e:
                logger.warning("Skip %s row %d: %s", sheet_name, row.idx, e)
                continue
            if not entity:
                continue
            db.add(ImportedEntity(
                batch_id=batch.id,
                entity_type=entity.__class__.__name__.lower(),
                entity_id=entity.id,
            ))
            _bump(sheet_name.lower())

    summary_bits = [f"{n} {kind}" for kind, n in counters.items() if n]
    batch.summary = " · ".join(summary_bits) or "Import vide"
    db.commit()
    db.refresh(batch)
    return batch


def _apply_resolutions(row: ParsedRow, sheet_res: Optional[dict]) -> dict:
    data = dict(row.data)
    if not sheet_res:
        return data
    patch = sheet_res.get(str(row.idx)) or sheet_res.get(row.idx)
    if patch:
        data.update(patch)
    return data


def _process_income(db, user, data, resolve_acc) -> Optional[Income]:
    acc = resolve_acc(data.get("account_name"))
    if not acc or not user_can_write_account(db, user.id, acc.id):
        return None
    try:
        itype = IncomeType(data.get("type") or "Régulier")
    except ValueError:
        itype = IncomeType.REGULIER
    i = Income(
        user_id=user.id,
        source=(data.get("source") or "").strip(),
        amount=_safe_decimal(data["amount"]) or Decimal(0),
        day_of_month=_safe_int(data["day_of_month"]) or 1,
        type=itype,
        account_id=acc.id,
        is_active=True,
        valid_from=_safe_date(data.get("valid_from")),
        valid_to=_safe_date(data.get("valid_to")),
    )
    db.add(i); db.flush(); return i


def _process_charge(db, user, data, resolve_acc) -> Optional[Charge]:
    acc = resolve_acc(data.get("account_name"))
    if not acc or not user_can_write_account(db, user.id, acc.id):
        return None
    try:
        freq = Frequency(data.get("frequency") or "Mensuelle")
    except ValueError:
        freq = Frequency.MENSUELLE
    try:
        sm = SplitMode(data.get("split_mode") or "Perso")
    except ValueError:
        sm = SplitMode.PERSO
    c = Charge(
        user_id=user.id,
        label=(data.get("label") or "").strip(),
        total_amount=_safe_decimal(data["total_amount"]) or Decimal(0),
        day_of_month=_safe_int(data["day_of_month"]) or 1,
        frequency=freq,
        split_mode=sm,
        num_colocs=_safe_int(data.get("num_colocs")) or 1,
        account_id=acc.id,
        is_active=True,
        valid_from=_safe_date(data.get("valid_from")),
        valid_to=_safe_date(data.get("valid_to")),
        notes=data.get("notes") or None,
    )
    db.add(c); db.flush(); return c


def _process_rec_transfer(db, user, data, resolve_acc) -> Optional[RecurringTransfer]:
    src = resolve_acc(data.get("source_account_name"))
    dst = resolve_acc(data.get("dest_account_name"))
    if not src or not dst:
        return None
    if not user_can_write_account(db, user.id, src.id) and not user_can_write_account(db, user.id, dst.id):
        return None
    try:
        freq = Frequency(data.get("frequency") or "Mensuelle")
    except ValueError:
        freq = Frequency.MENSUELLE
    t = RecurringTransfer(
        user_id=user.id,
        label=(data.get("label") or "").strip(),
        amount=_safe_decimal(data["amount"]) or Decimal(0),
        source_account_id=src.id,
        dest_account_id=dst.id,
        day_of_month=_safe_int(data["day_of_month"]) or 1,
        frequency=freq,
        is_active=True,
        valid_from=_safe_date(data.get("valid_from")),
        valid_to=_safe_date(data.get("valid_to")),
    )
    db.add(t); db.flush(); return t


def _process_ot_transfer(db, user, data, resolve_acc) -> Optional[OneTimeTransfer]:
    src = resolve_acc(data.get("source_account_name"))
    dst = resolve_acc(data.get("dest_account_name"))
    if not src or not dst:
        return None
    if not user_can_write_account(db, user.id, src.id) and not user_can_write_account(db, user.id, dst.id):
        return None
    t = OneTimeTransfer(
        user_id=user.id,
        date=_safe_date(data["date"]) or DateType.today(),
        label=(data.get("label") or "").strip(),
        amount=_safe_decimal(data["amount"]) or Decimal(0),
        source_account_id=src.id,
        dest_account_id=dst.id,
    )
    db.add(t); db.flush(); return t


def _process_saving(db, user, data, resolve_acc) -> Optional[AutoSaving]:
    src = resolve_acc(data.get("source_account_name"))
    dst = resolve_acc(data.get("dest_account_name"))
    if not src or not dst:
        return None
    if not user_can_write_account(db, user.id, src.id) and not user_can_write_account(db, user.id, dst.id):
        return None
    s = AutoSaving(
        user_id=user.id,
        label=(data.get("label") or "").strip(),
        amount=_safe_decimal(data["amount"]) or Decimal(0),
        source_account_id=src.id,
        dest_account_id=dst.id,
        day_of_month=_safe_int(data["day_of_month"]) or 1,
        is_active=True,
        valid_from=_safe_date(data.get("valid_from")),
        valid_to=_safe_date(data.get("valid_to")),
    )
    db.add(s); db.flush(); return s


def _process_purchase(db, user, data, resolve_acc) -> Optional[Purchase]:
    acc = resolve_acc(data.get("account_name"))
    if not acc or not user_can_write_account(db, user.id, acc.id):
        return None
    try:
        pm = PaymentMethod(data.get("payment_method") or "CB")
    except ValueError:
        pm = PaymentMethod.CB
    p = Purchase(
        user_id=user.id,
        date=_safe_date(data["date"]) or DateType.today(),
        description=(data.get("description") or "").strip(),
        total_amount=_safe_decimal(data["total_amount"]) or Decimal(0),
        nb_installments=_safe_int(data.get("nb_installments")) or 1,
        category=data.get("category") or None,
        payment_method=pm,
        account_id=acc.id,
    )
    db.add(p); db.flush(); return p
